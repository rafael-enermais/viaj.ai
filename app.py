# Viaj.AI — v3.0 (Assistente com ACAO: propor lancamento rapido por linguagem natural, grava so' apos confirmacao humana) — ver 00-handoff.md do VIAJAI no vault
# Gestão de folgas, deslocamento e custo de funcionários em obra — EnerMais.
#
# Reaproveita o padrão validado em produção do TIA.go/RHDADOS:
# - cliente Supabase sempre em st.session_state (nunca st.cache_resource)
# - reanexar token a cada rerun (supabase.postgrest.auth(token))
# - login só Supabase Auth (e-mail/senha)
#
# NAO RODA ainda sem: (1) schema `viajai` v0.1 a v0.4 rodados no Supabase
# (ver arquivos/schema_v0.*.txt no vault) e (2) .env preenchido a partir
# de .env.example.
#
# PARSER DO RE090 — NOTA IMPORTANTE (01/09): as 10 planilhas que seguem o
# padrão RE090 NAO tem a mesma ordem de coluna entre si (confirmado lendo
# 4 arquivos reais: uma tem uma coluna extra "CONFERENCIA RH" antes do
# nome, deslocando tudo; duas nao tem as colunas "ultima folga"/"qnt dias"
# de jeito nenhum). Por isso o parser abaixo NUNCA usa indice fixo de
# coluna — ele le a linha de cabecalho (linha 4, confirmado em 4/4 arquivos
# reais) e casa pelo TEXTO do cabecalho. Se um arquivo novo tiver cabecalho
# diferente o suficiente pra nao casar "nome", a tela avisa em vez de
# importar linha errada silenciosamente.

import base64
import io
import os
from datetime import date, datetime

import anthropic
import openpyxl
import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from supabase import create_client, Client

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")

# TODO (03/09): confirmado contra a doc oficial da Anthropic em 02/09/2026
# que "claude-sonnet-5" e' o identificador de modelo atual - MAS confirma
# de novo aqui se isso mudar (mesmo erro que o TIA.go deixou marcado: nunca
# hardcoded sem checar contra a conta/doc na hora).
MODEL_ID = "claude-sonnet-5"

st.set_page_config(page_title="Viaj.AI", page_icon="🧳", layout="wide")

LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo-enermais.png")


@st.cache_data
def logo_base64():
    with open(LOGO_PATH, "rb") as f:
        return base64.b64encode(f.read()).decode()


def render_logo(height=56):
    # Mesmo padrão do TIA.go/Radar/RHDADOS: logo oficial (navy+laranja) em
    # silhueta branca via filtro CSS (brightness(0) invert(1)), sem precisar
    # de PNG separado. Tema base agora e' "dark" (.streamlit/config.toml),
    # entao usa o mesmo filtro do TIA.go (fundo escuro -> silhueta branca).
    st.markdown(
        f"""
        <img src="data:image/png;base64,{logo_base64()}" height="{height}"
             style="filter: brightness(0) invert(1); margin-bottom: 12px;">
        """,
        unsafe_allow_html=True,
    )

ABA_FOLGAS = "Controle de Folgas"
LINHA_CABECALHO = 4  # confirmado em 4/4 arquivos RE090 reais analisados

# texto do cabecalho (minusculo, sem acento no essencial) -> chave interna
ALIASES_COLUNA = {
    "centro de custo": "centro_custo",
    "nome do funcion": "nome",  # match parcial: "Nome do Funcionário (completo)"
    "ultima folga": "ultima_folga",
    "última folga": "ultima_folga",
    "inicio da folga": "inicio_folga",
    "início da folga": "inicio_folga",
    "termino da folga": "termino_folga",
    "término da folga": "termino_folga",
}

# Codigo IATA por cidade (fato estavel de aviacao, nao preco/horario - sem
# risco de ficar desatualizado do jeito que preco fica). So' as capitais/
# hubs mais comuns; se a cidade digitada nao estiver aqui, usuario digita
# o codigo direto (ex.: "FOR") que a funcao abaixo aceita igual.
_IATA_CIDADES = {
    "fortaleza": "FOR", "sao paulo": "GRU", "são paulo": "GRU",
    "rio de janeiro": "GIG", "salvador": "SSA", "recife": "REC",
    "brasilia": "BSB", "brasília": "BSB", "belo horizonte": "CNF",
    "curitiba": "CWB", "porto alegre": "POA", "manaus": "MAO",
    "belem": "BEL", "belém": "BEL", "vitoria": "VIX", "vitória": "VIX",
    "natal": "NAT", "joao pessoa": "JPA", "joão pessoa": "JPA",
    "maceio": "MCZ", "maceió": "MCZ", "aracaju": "AJU", "teresina": "THE",
    "sao luis": "SLZ", "são luís": "SLZ", "cuiaba": "CGB", "cuiabá": "CGB",
    "campo grande": "CGR", "goiania": "GYN", "goiânia": "GYN",
    "florianopolis": "FLN", "florianópolis": "FLN",
}


def _resolver_iata(texto):
    """Aceita nome de cidade (casa pelo dicionario acima) ou codigo IATA
    direto (3 letras) - devolve o codigo em maiusculo, ou string vazia se
    nao reconhecer nem uma coisa nem outra (usuario ve e corrige)."""
    if not texto:
        return ""
    limpo = texto.strip()
    if len(limpo) == 3 and limpo.isalpha():
        return limpo.upper()
    return _IATA_CIDADES.get(limpo.lower(), "")


def _botao_exportar_excel(df, nome_arquivo, label="Exportar Excel"):
    """Exporta o df atual (como esta na tela, ja filtrado/ordenado) pra .xlsx.
    Pedido do Rafael (02/09): dar pra levar a tabela pra fora do app, ex.
    compartilhar com gestor de obra que nao usa o Viaj.AI."""
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    st.download_button(
        label,
        data=buffer.getvalue(),
        file_name=nome_arquivo,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"export_{nome_arquivo}",
    )


def get_client() -> Client:
    if "supabase_client" not in st.session_state:
        st.session_state.supabase_client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
    return st.session_state.supabase_client


def tela_login():
    render_logo(height=64)
    st.caption("Viaj.AI — Gestão de folgas, deslocamento e custo")
    with st.form("login"):
        email = st.text_input("E-mail")
        senha = st.text_input("Senha", type="password")
        entrar = st.form_submit_button("Entrar")
    if entrar:
        supabase = get_client()
        try:
            resp = supabase.auth.sign_in_with_password({"email": email, "password": senha})
            st.session_state.sessao = resp.session
            st.session_state.usuario = email
            st.rerun()
        except Exception as e:
            st.error(f"Login falhou: {e}")


def _mapear_colunas(ws, linha_cabecalho=LINHA_CABECALHO, max_colunas=25):
    """Le a linha de cabecalho e devolve {chave_interna: indice_coluna(1-based)}.
    Casamento por trecho de texto, nao por posicao — planilhas reais tem
    colunas em ordens diferentes (ver nota no topo do arquivo)."""
    mapa = {}
    for c in range(1, max_colunas + 1):
        valor = ws.cell(row=linha_cabecalho, column=c).value
        if not valor:
            continue
        texto = str(valor).strip().lower()
        for trecho, chave in ALIASES_COLUNA.items():
            if trecho in texto and chave not in mapa:
                mapa[chave] = c
    return mapa


def _valor_data(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    return None  # texto solto em campo de data: nao adivinha, descarta


def _ler_planilha_re090(arquivo):
    # read_only=True é obrigatório aqui: testado contra os 14 arquivos reais
    # e o modo padrão (carrega tudo na memória, incl. imagem/formatação)
    # trava/mata o processo no arquivo de 19MB (NATANAEL). Com read_only
    # esse mesmo arquivo leu em 0.1s.
    wb = openpyxl.load_workbook(arquivo, data_only=True, read_only=True)
    if ABA_FOLGAS not in wb.sheetnames:
        return None, f"Aba '{ABA_FOLGAS}' não encontrada. Abas nesse arquivo: {wb.sheetnames}"

    ws = wb[ABA_FOLGAS]
    mapa = _mapear_colunas(ws)
    if "nome" not in mapa:
        return None, (
            "Não achei a coluna de nome do funcionário no cabeçalho da linha "
            f"{LINHA_CABECALHO}. Essa planilha pode não seguir o padrão RE090 "
            "esperado — confere manualmente antes de tentar de novo."
        )

    linhas = []
    for row in ws.iter_rows(min_row=LINHA_CABECALHO + 1, values_only=True):
        def pega(chave):
            idx = mapa.get(chave)
            return row[idx - 1] if idx and idx - 1 < len(row) else None

        nome = pega("nome")
        if not nome or not str(nome).strip():
            continue

        linhas.append({
            "nome": str(nome).strip(),
            "centro_custo": pega("centro_custo"),
            "data_ultimo_retorno": _valor_data(pega("ultima_folga")),
            "data_saida_prevista": _valor_data(pega("inicio_folga")),
            "data_retorno_prevista": _valor_data(pega("termino_folga")),
        })
    return linhas, None


def _importar_um_arquivo(supabase, arquivo):
    """Processa 1 arquivo RE090: le, abre um lote (import_batch), roda a RPC
    linha a linha carimbando o lote, fecha o lote com os totais. Devolve
    (nome_arquivo, resultados, erro_leitura)."""
    linhas, erro = _ler_planilha_re090(arquivo)
    if erro:
        return arquivo.name, None, erro

    batch = supabase.rpc("viajai_criar_import_batch", {
        "p_nome_arquivo": arquivo.name,
        "p_total_linhas": len(linhas),
    }).execute()
    batch_id = batch.data

    resultados = []
    for l in linhas:
        try:
            resp = supabase.rpc("viajai_importar_re090_linha", {
                "p_nome_planilha": l["nome"],
                "p_matricula_planilha": None,
                "p_texto_obra_canteiro": l["centro_custo"],
                "p_data_ultimo_retorno": l["data_ultimo_retorno"].isoformat() if l["data_ultimo_retorno"] else None,
                "p_data_saida_prevista": l["data_saida_prevista"].isoformat() if l["data_saida_prevista"] else None,
                "p_data_retorno_prevista": l["data_retorno_prevista"].isoformat() if l["data_retorno_prevista"] else None,
                "p_import_batch_id": batch_id,
            }).execute()
            linha_resultado = resp.data[0] if resp.data else {"resultado": "sem_retorno"}
        except Exception as e:
            linha_resultado = {"resultado": "erro", "colaborador_id": None, "folga_id": None}
            st.warning(f"Erro ao importar '{l['nome']}' ({arquivo.name}): {e}")
        resultados.append({"nome": l["nome"], **linha_resultado})

    criadas = sum(1 for r in resultados if r.get("resultado") == "criada")
    pendentes = sum(1 for r in resultados if r.get("resultado") == "pendencia")
    supabase.rpc("viajai_finalizar_import_batch", {
        "p_batch_id": batch_id,
        "p_total_criadas": criadas,
        "p_total_pendencias": pendentes,
    }).execute()

    return arquivo.name, resultados, None


def pagina_importar_re090(supabase):
    st.subheader("Importar RE090")
    st.caption(
        "Sobe a(s) planilha(s), resolve cada colaborador contra o RH (ao vivo) "
        "e grava a folga já com o canteiro espelho. Sem match único, a linha "
        "vira pendência — nada é perdido, só fica pra você revisar. Pode subir "
        "mais de um arquivo de uma vez — cada arquivo vira 1 lote no histórico, "
        "revertível separadamente."
    )

    arquivos = st.file_uploader(
        "Planilha(s) RE090 (.xlsx)", type=["xlsx"], accept_multiple_files=True
    )
    if arquivos:
        pre_leituras = []
        for arquivo in arquivos:
            linhas, erro = _ler_planilha_re090(arquivo)
            pre_leituras.append((arquivo, linhas, erro))
            if erro:
                st.error(f"{arquivo.name}: {erro}")
            else:
                st.write(f"**{arquivo.name}** — {len(linhas)} linha(s) com nome preenchido.")
                with st.expander(f"Ver linhas antes de importar — {arquivo.name}"):
                    st.dataframe(linhas)

        if st.button("Importar", type="primary"):
            resultado_por_arquivo = {}
            for arquivo, linhas, erro in pre_leituras:
                if erro:
                    continue
                nome, resultados, _ = _importar_um_arquivo(supabase, arquivo)
                resultado_por_arquivo[nome] = resultados
            st.session_state["resultado_import"] = resultado_por_arquivo

    if "resultado_import" in st.session_state:
        for nome_arquivo, resultados in st.session_state["resultado_import"].items():
            criadas = sum(1 for r in resultados if r.get("resultado") == "criada")
            pendentes = sum(1 for r in resultados if r.get("resultado") == "pendencia")
            duplicadas = sum(1 for r in resultados if r.get("resultado") == "duplicada")
            msg = f"**{nome_arquivo}**: {criadas} folga(s) criada(s), {pendentes} em pendência"
            if duplicadas:
                msg += f", {duplicadas} duplicada(s) (já existia folga prevista pra essa pessoa nessa mesma data — ignorada, não criou de novo)"
            st.success(msg + ".")
            st.dataframe(resultados)

    st.divider()
    st.subheader("Histórico de imports")
    st.caption(
        "Cada linha é 1 upload. Reverter apaga as folgas criadas por esse "
        "lote específico — só as que ninguém mexeu ainda (status ainda "
        "'prevista', sem trecho de viagem associado); o resto fica registrado "
        "mas não é apagado, pra não perder trabalho já feito em cima."
    )
    lotes = supabase.rpc("viajai_listar_import_batches", {"p_limite": 20}).execute()
    if lotes.data:
        for lote in lotes.data:
            cols = st.columns([3, 2, 2, 2, 2])
            cols[0].write(f"{lote['nome_arquivo']}")
            cols[1].write(lote["criado_em"][:16].replace("T", " "))
            cols[2].write(f"{lote['total_criadas']} criada(s)")
            cols[3].write(f"{lote['total_pendencias']} pendência(s)")
            if lote["revertido"]:
                cols[4].write("↩️ revertido")
            else:
                if cols[4].button("Reverter", key=f"reverter_{lote['id']}"):
                    rev = supabase.rpc(
                        "viajai_reverter_import_batch", {"p_batch_id": lote["id"]}
                    ).execute()
                    r = rev.data[0] if rev.data else {}
                    st.success(
                        f"Revertido: {r.get('folgas_removidas', 0)} folga(s) removida(s), "
                        f"{r.get('folgas_puladas', 0)} pulada(s) (já tinham sido mexidas), "
                        f"{r.get('pendencias_removidas', 0)} pendência(s) removida(s)."
                    )
                    st.rerun()
    else:
        st.caption("Nenhum import feito ainda.")

    st.divider()
    st.subheader("Pendências abertas")
    st.caption(
        "Marca 'Resolvida' pra tirar da lista (revisou manualmente e não "
        "precisa mais aparecer aqui — não cria folga nenhuma, só limpa a fila)."
    )
    pend = supabase.rpc("viajai_listar_pendencias_import", {"p_apenas_nao_resolvidas": True}).execute()
    if pend.data:
        df_pend = pd.DataFrame(pend.data)
        df_pend = df_pend.drop(columns=[c for c in ["origem"] if c in df_pend.columns])
        df_pend["resolvido"] = False
        editado_pend = st.data_editor(
            df_pend,
            column_config={
                "resolvido": st.column_config.CheckboxColumn(
                    "Resolvida", help="Marca e clica em Salvar embaixo pra tirar da lista."
                ),
            },
            disabled=[c for c in df_pend.columns if c not in ("resolvido",)],
            hide_index=True,
            use_container_width=True,
            key="editor_pendencias",
        )
        _botao_exportar_excel(df_pend.drop(columns=["resolvido"]), "viajai_pendencias.xlsx")
        if st.button("Salvar pendências resolvidas"):
            marcadas = editado_pend[editado_pend["resolvido"] == True]  # noqa: E712
            if marcadas.empty:
                st.info("Nenhuma pendência marcada — nada pra salvar.")
            else:
                erros = 0
                for _, linha in marcadas.iterrows():
                    try:
                        supabase.rpc("viajai_marcar_pendencia_resolvida", {
                            "p_pendencia_id": int(linha["id"]),
                            "p_resolvido": True,
                        }).execute()
                    except Exception as e:
                        erros += 1
                        st.error(f"Erro ao resolver pendência {linha['id']}: {e}")
                sucesso = len(marcadas) - erros
                if sucesso:
                    st.success(f"{sucesso} pendência(s) marcada(s) como resolvida(s).")
                st.rerun()
    else:
        st.caption("Nenhuma pendência em aberto.")


_STATUS_OPCOES = ["prevista", "confirmada", "em_andamento", "realizada", "vendida", "cancelada"]


def pagina_confirmar_folgas(supabase):
    st.subheader("Confirmar folgas")
    st.caption(
        "Fecha o ciclo: registre aqui o que realmente aconteceu com cada "
        "folga que caiu como 'prevista' (via import ou manual). Sem isso, "
        "a Previsão de folgas nunca reflete a realidade — só o que foi "
        "planejado. Edita direto na tabela: muda o 'Status novo' de quem "
        "mudou, preenche data real se for o caso, e clica em Salvar no "
        "final — quem ficar em 'prevista' não é tocado."
    )

    previstas = supabase.rpc("viajai_listar_folgas_previstas", {"p_limite": 300}).execute()
    if not previstas.data:
        st.caption("Nenhuma folga 'prevista' aguardando confirmação no momento.")
    else:
        base = pd.DataFrame(previstas.data)
        # "atrasada" = data de saida prevista ja passou e ninguem confirmou
        # nada ainda (pedido do Rafael: risco real de colaborador seguir na
        # obra alem do previsto sem registro). So' compara com hoje, nao
        # precisa de RPC nova.
        _hoje = date.today()
        base["data_saida_prevista"] = pd.to_datetime(base["data_saida_prevista"]).dt.date
        base["situacao"] = base["data_saida_prevista"].apply(
            lambda d: "⚠️ atrasada" if pd.notna(d) and d < _hoje else "no prazo"
        )

        # urgencia (mesmo calculo da tela Previsao de folgas) trazida por
        # colaborador_id - pedido do Rafael: ajuda a Amanda a priorizar
        # quem confirmar primeiro sem precisar trocar de tela.
        prev_resp = supabase.rpc("viajai_previsao_folgas").execute()
        if prev_resp.data:
            df_urg = pd.DataFrame(prev_resp.data)[["colaborador_id", "nivel_urgencia"]]
            df_urg = df_urg.rename(columns={"nivel_urgencia": "urgencia"})
            base = base.merge(df_urg, on="colaborador_id", how="left")
        else:
            base["urgencia"] = None
        base["urgencia"] = base["urgencia"].fillna("—")
        base = base.sort_values(by=["situacao", "data_saida_prevista"], ascending=[True, True])
        base["status_novo"] = "prevista"
        base["data_saida_real"] = pd.NaT
        base["data_retorno_real"] = pd.NaT
        base["motivo_venda"] = ""

        editado = st.data_editor(
            base,
            column_order=[
                "situacao", "urgencia", "nome", "obra_nome", "canteiro_nome",
                "data_saida_prevista", "data_retorno_prevista",
                "status_novo", "data_saida_real", "data_retorno_real", "motivo_venda",
            ],
            column_config={
                "situacao": st.column_config.TextColumn("Situação", disabled=True),
                "urgencia": st.column_config.TextColumn(
                    "Urgência", disabled=True,
                    help="Mesmo cálculo da tela Previsão de folgas — pra editar o override, vai lá.",
                ),
                "nome": st.column_config.TextColumn("Nome", disabled=True),
                "obra_nome": st.column_config.TextColumn("Obra", disabled=True),
                "canteiro_nome": st.column_config.TextColumn("Canteiro", disabled=True),
                "data_saida_prevista": st.column_config.DateColumn("Saída prevista", disabled=True),
                "data_retorno_prevista": st.column_config.DateColumn("Retorno previsto", disabled=True),
                "status_novo": st.column_config.SelectboxColumn(
                    "Status novo", options=_STATUS_OPCOES, required=True,
                    help="Deixa 'prevista' pra não mexer nessa linha.",
                ),
                "data_saida_real": st.column_config.DateColumn(
                    "Saída real", help="Preenche se marcou 'em_andamento' ou 'realizada'."
                ),
                "data_retorno_real": st.column_config.DateColumn(
                    "Retorno real", help="Preenche se marcou 'realizada'."
                ),
                "motivo_venda": st.column_config.TextColumn(
                    "Motivo (se vendida)", help="Opcional, só faz sentido se marcou 'vendida'."
                ),
            },
            hide_index=True,
            use_container_width=True,
            key="editor_confirmar_folgas",
        )
        _botao_exportar_excel(
            base.drop(columns=["status_novo", "data_saida_real", "data_retorno_real", "motivo_venda"]),
            "viajai_confirmar_folgas.xlsx",
        )
        st.caption("'Urgência' é só leitura aqui — pra travar manual (override), usa a tela 'Previsão de folgas'.")

        if st.button("Salvar alterações", type="primary"):
            mudou = editado[editado["status_novo"] != "prevista"]
            if mudou.empty:
                st.info("Nenhuma linha teve o status alterado — nada pra salvar.")
            else:
                erros = 0
                for _, linha in mudou.iterrows():
                    try:
                        supabase.rpc("viajai_atualizar_folga", {
                            "p_folga_id": int(linha["folga_id"]),
                            "p_status": linha["status_novo"],
                            "p_data_saida_real": (
                                linha["data_saida_real"].isoformat()
                                if pd.notna(linha["data_saida_real"]) else None
                            ),
                            "p_data_retorno_real": (
                                linha["data_retorno_real"].isoformat()
                                if pd.notna(linha["data_retorno_real"]) else None
                            ),
                            "p_motivo_venda": linha["motivo_venda"] or None,
                        }).execute()
                    except Exception as e:
                        erros += 1
                        st.error(f"Erro ao salvar {linha['nome']}: {e}")
                sucesso = len(mudou) - erros
                if sucesso:
                    st.success(f"{sucesso} folga(s) atualizada(s).")
                st.rerun()

        st.caption(f"{len(previstas.data)} folga(s) prevista(s) aguardando confirmação.")

    st.divider()
    st.subheader("Histórico")
    st.caption("Últimas mudanças registradas — quem, quando, o que mudou.")
    hist = supabase.rpc("viajai_listar_historico_folga", {"p_limite": 100}).execute()
    if hist.data:
        st.dataframe(hist.data, use_container_width=True, hide_index=True)
    else:
        st.caption("Nenhuma mudança registrada ainda.")


_ORDEM_URGENCIA = {"atrasado": -1, "critico": 0, "atencao": 1, "normal": 2}


def pagina_previsao(supabase):
    st.subheader("Previsão de folgas")
    st.caption(
        "Calculado ao vivo: colaborador ativo no RH + última folga conhecida "
        "no Viaj.AI. Sem histórico ainda = sem previsão (precisa de ao menos "
        "1 folga registrada, manual ou via import)."
    )
    resp = supabase.rpc("viajai_previsao_folgas").execute()
    if not resp.data:
        st.caption("Sem dados ainda.")
        return

    df = pd.DataFrame(resp.data)

    # previsao de gasto por colaborador (schema_v0.15): media do proprio
    # historico, cai pra media do canteiro, depois da obra, se nao tiver -
    # pedido do Rafael (02/09) pra dar valor em R$ nessa tabela tambem.
    gasto_resp = supabase.rpc("viajai_previsao_gasto_colaborador").execute()
    if gasto_resp.data:
        df_gasto = pd.DataFrame(gasto_resp.data)
        df = df.merge(df_gasto, on="colaborador_id", how="left")
    else:
        df["previsao_gasto"] = None
        df["base_previsao"] = "sem_dado"

    # painel resumido (pedido do Rafael: "dimensionar quantos funcionarios
    # precisarao de passagem nos proximos 30 dias" + gasto previsto) - usa
    # o df ANTES do filtro de "mostrar sem historico" (precisa de
    # dias_restantes calculado, que so existe pra quem tem historico).
    _janela_30d = df[
        df["dias_restantes"].notna()
        & (df["dias_restantes"] >= 0)
        & (df["dias_restantes"] <= 30)
    ]
    _qtd_30d = len(_janela_30d)
    _com_previsao = _janela_30d["previsao_gasto"].notna().sum() if _qtd_30d else 0
    _gasto_30d = _janela_30d["previsao_gasto"].sum(skipna=True) if _qtd_30d else 0
    kpi1, kpi2 = st.columns(2)
    kpi1.metric("Precisando de passagem (próx. 30 dias)", _qtd_30d)
    if _com_previsao:
        kpi2.metric(
            "Gasto previsto pra esse período",
            f"R$ {_gasto_30d:,.2f}",
            help=f"Baseado em {_com_previsao} de {_qtd_30d} pessoa(s) com histórico de custo — o resto ainda não tem base pra estimar.",
        )
    else:
        kpi2.metric("Gasto previsto pra esse período", "sem dado ainda")
    st.divider()

    mostrar_sem_historico = st.checkbox(
        "Mostrar também quem ainda não tem nenhuma folga registrada "
        "(sem previsão calculável ainda)",
        value=False,
    )
    if not mostrar_sem_historico:
        df = df[df["tem_historico"] == True]  # noqa: E712
        if df.empty:
            st.caption(
                "Ninguém com histórico ainda — importe pelo menos 1 RE090 ou "
                "marque a caixa acima pra ver a lista completa sem previsão."
            )
            return

    # "atrasado" = dias_restantes negativo (a data prevista de saida ja
    # passou e ninguem registrou nada ainda) - risco real, pedido do Rafael.
    # So' compara com o que ja vem calculado, nao precisa de RPC nova.
    df["situacao"] = df.apply(
        lambda r: "atrasado" if pd.notna(r["dias_restantes"]) and r["dias_restantes"] < 0 else r["nivel_urgencia"],
        axis=1,
    )

    # mais urgente primeiro: atrasado > critico > atencao > normal > (sem classificacao)
    # e, dentro do mesmo nivel, quem tem menos dias restantes primeiro
    df["_ordem_urgencia"] = df["situacao"].map(_ORDEM_URGENCIA).fillna(9)
    df = df.sort_values(by=["_ordem_urgencia", "dias_restantes"], na_position="last")
    df = df.drop(columns=["_ordem_urgencia"])

    # override manual de urgencia: nivel_urgencia ja vem calculado com o
    # override aplicado (COALESCE no RPC) e urgencia_manual diz se existe
    # override pra aquela pessoa - entao da pra reconstruir o valor bruto
    # do override sem precisar de outra RPC.
    _UO_AUTOMATICO = "(automático)"
    df["override_manual"] = df.apply(
        lambda r: r["nivel_urgencia"] if r.get("urgencia_manual") else _UO_AUTOMATICO,
        axis=1,
    )

    colunas_principais = [
        "nome", "situacao", "dias_restantes",
        "data_saida_prevista", "data_retorno_prevista",
        "obra_nome", "canteiro_nome", "previsao_gasto", "base_previsao", "override_manual",
    ]
    colunas_principais = [c for c in colunas_principais if c in df.columns]
    # colunas tecnicas (ids, flags internas de ordenacao/filtro) ficam de
    # fora da visualizacao, mas continuam no df (usadas ao salvar).
    colunas_ocultas = {
        "obra_id", "canteiro_id", "data_base_retorno",
        "tem_historico", "urgencia_manual", "nivel_urgencia",
    }
    outras = [c for c in df.columns if c not in colunas_principais and c not in colunas_ocultas]
    df = df[colunas_principais + outras]

    st.caption(
        "'Urgência' (coluna calculada) muda sozinha conforme os dias passam. "
        "Pra travar manualmente pra 1 pessoa (ex.: sabe que ela vai atrasar "
        "por outro motivo), muda 'Override manual' e clica Salvar embaixo — "
        "'(automático)' volta a deixar o cálculo decidir sozinho."
    )
    editado = st.data_editor(
        df,
        column_config={
            "override_manual": st.column_config.SelectboxColumn(
                "Override manual",
                options=[_UO_AUTOMATICO, "critico", "atencao", "normal"],
                required=True,
            ),
            "previsao_gasto": st.column_config.NumberColumn(
                "Previsão de gasto",
                format="R$ %.2f",
                help="Média do custo histórico (da própria pessoa; sem isso, do canteiro; sem isso, da obra) — ver coluna 'Base'.",
            ),
            "base_previsao": st.column_config.TextColumn(
                "Base", help="De onde veio a previsão: colaborador, canteiro, obra ou sem_dado (nunca registrado nada ainda)."
            ),
        },
        disabled=[c for c in df.columns if c != "override_manual"],
        column_order=colunas_principais,
        hide_index=True,
        use_container_width=True,
        key="editor_previsao",
    )
    _botao_exportar_excel(df.drop(columns=["override_manual"]), "viajai_previsao_folgas.xlsx")

    if st.button("Salvar overrides de urgência"):
        mudou = editado[editado["override_manual"] != df["override_manual"]]
        if mudou.empty:
            st.info("Nenhum override mudou — nada pra salvar.")
        else:
            erros = 0
            for _, linha in mudou.iterrows():
                try:
                    if linha["override_manual"] == _UO_AUTOMATICO:
                        supabase.rpc("viajai_remover_urgencia_override", {
                            "p_colaborador_id": linha["colaborador_id"],
                        }).execute()
                    else:
                        supabase.rpc("viajai_definir_urgencia_override", {
                            "p_colaborador_id": linha["colaborador_id"],
                            "p_nivel": linha["override_manual"],
                        }).execute()
                except Exception as e:
                    erros += 1
                    st.error(f"Erro ao salvar override de {linha['nome']}: {e}")
            sucesso = len(mudou) - erros
            if sucesso:
                st.success(f"{sucesso} override(s) salvo(s).")
            st.rerun()

    st.divider()
    st.subheader("Desvio de planejamento")
    st.caption(
        "Prevista x real: só datas por enquanto (custo em R$ ainda não tem "
        "tela pra registrar — viagem/trecho/gasto é o próximo bloco grande, "
        "ver 00-handoff). Positivo = atrasou em relação ao previsto; "
        "negativo = antecipou."
    )
    desvio = supabase.rpc("viajai_listar_folgas_desvio", {"p_limite": 200}).execute()
    if desvio.data:
        df_desvio = pd.DataFrame(desvio.data)
        st.dataframe(df_desvio, use_container_width=True, hide_index=True)
        _botao_exportar_excel(df_desvio, "viajai_desvio_planejamento.xlsx")
    else:
        st.caption("Nenhuma folga confirmada/realizada ainda pra comparar.")



def pagina_custo_passagens(supabase):
    st.subheader("Custo & Passagens")
    st.caption(
        "Registra o preço REAL pago (passagem/ônibus/carro) e gastos extras "
        "por folga, ou um lançamento rápido solto quando não dá pra apontar "
        "folga específica na hora. Sem busca de preço ao vivo (decisão "
        "02/09 — ver 00-handoff): a inteligência aqui é o histórico que a "
        "própria Amanda for alimentando, cresce com o uso."
    )

    with st.expander("🔎 Consultar histórico da rota (antes de comprar)"):
        st.caption(
            "Usa o que já foi registrado no Viaj.AI — escreva origem/destino "
            "igual a como costuma registrar (mesmo texto), senão não casa."
        )
        c1, c2 = st.columns(2)
        origem_c = c1.text_input("Origem", key="consulta_origem")
        destino_c = c2.text_input("Destino", key="consulta_destino")
        if st.button("Consultar histórico", key="btn_consulta_hist"):
            if origem_c and destino_c:
                sug = supabase.rpc("viajai_sugestao_fornecedor_rota", {
                    "p_origem": origem_c, "p_destino": destino_c,
                }).execute()
                if sug.data:
                    st.write("**Fornecedor mais usado nessa rota:**")
                    st.dataframe(sug.data, hide_index=True, use_container_width=True)
                else:
                    st.caption("Sem fornecedor registrado ainda pra essa rota.")

                comp_modal = supabase.rpc("viajai_comparar_modais_rota", {
                    "p_origem": origem_c, "p_destino": destino_c,
                }).execute()
                if comp_modal.data:
                    st.write("**Comparativo por modal (preço médio e duração média):**")
                    st.dataframe(comp_modal.data, hide_index=True, use_container_width=True)
                else:
                    st.caption("Sem dado suficiente ainda pra comparar modal nessa rota.")
            else:
                st.info("Preenche origem e destino.")

        st.divider()
        st.caption(
            "Atalho pra abrir busca já preenchida no Skyscanner (link oficial "
            "deles, sem API key) — digita nome da cidade (ex.: Fortaleza) ou "
            "já o código do aeroporto (ex.: FOR), os dois funcionam:"
        )
        c3, c4, c5 = st.columns(3)
        origem_txt = c3.text_input("Origem", key="sky_origem")
        destino_txt = c4.text_input("Destino", key="sky_destino")
        data_ida_sky = c5.date_input("Data de ida", value=date.today(), key="sky_data")
        origem_iata = _resolver_iata(origem_txt)
        destino_iata = _resolver_iata(destino_txt)
        if origem_txt and not origem_iata:
            st.caption(f"Não reconheci '{origem_txt}' — digita o código do aeroporto direto (ex.: FOR).")
        if destino_txt and not destino_iata:
            st.caption(f"Não reconheci '{destino_txt}' — digita o código do aeroporto direto (ex.: FOR).")
        if origem_iata and destino_iata:
            url_sky = (
                "https://www.skyscanner.net/g/referrals/v1/flights/day-view/"
                f"?origin={origem_iata}&destination={destino_iata}"
                f"&outboundDate={data_ida_sky.isoformat()}&market=BR&currency=BRL&locale=pt-BR"
            )
            st.link_button(f"🔗 Ver no Skyscanner ({origem_iata} → {destino_iata})", url_sky)
        else:
            st.caption("Preenche origem e destino (cidade ou código) pra habilitar o link.")

    aba_folga, aba_rapido = st.tabs(["Por folga", "Lançamento rápido"])

    with aba_folga:
        folgas_resp = supabase.rpc("viajai_listar_folgas_desvio", {"p_limite": 500}).execute()
        if not folgas_resp.data:
            st.caption(
                "Nenhuma folga confirmada/realizada ainda — confirme pelo "
                "menos 1 na tela 'Confirmar folgas' antes de registrar custo."
            )
        else:
            df_folgas = pd.DataFrame(folgas_resp.data)
            df_folgas["rotulo"] = df_folgas.apply(
                lambda r: f"#{r['folga_id']} — {r['nome']} — {r.get('canteiro_nome') or '—'} — {r['status']}",
                axis=1,
            )
            escolha = st.selectbox("Folga", df_folgas["rotulo"], key="folga_custo_select")
            folga_id_sel = int(df_folgas.loc[df_folgas["rotulo"] == escolha, "folga_id"].iloc[0])

            viagens_resp = supabase.rpc("viajai_listar_viagens_folga", {"p_folga_id": folga_id_sel}).execute()
            st.write("**Trechos já registrados:**")
            if viagens_resp.data:
                st.dataframe(pd.DataFrame(viagens_resp.data), hide_index=True, use_container_width=True)
            else:
                st.caption("Nenhum trecho registrado ainda pra essa folga.")

            gastos_resp = supabase.rpc("viajai_listar_gastos_folga", {"p_folga_id": folga_id_sel}).execute()
            st.write("**Gastos extras já registrados:**")
            if gastos_resp.data:
                st.dataframe(pd.DataFrame(gastos_resp.data), hide_index=True, use_container_width=True)
            else:
                st.caption("Nenhum gasto extra registrado ainda.")

            with st.form("form_add_trecho"):
                st.write("Adicionar trecho (passagem/perna da viagem)")
                st.caption("Só o essencial aqui — o resto é opcional, fica em 'Mais detalhes'.")
                c1, c2 = st.columns(2)
                sentido = c1.selectbox("Sentido", ["ida", "volta"])
                modal = c2.selectbox("Modal", ["aviao", "onibus", "carro", "taxi", "outro"])
                c3, c4 = st.columns(2)
                origem_t = c3.text_input("Origem", key="trecho_origem")
                destino_t = c4.text_input("Destino", key="trecho_destino")
                c5, c6 = st.columns(2)
                preco_t = c5.number_input("Preço (R$)", min_value=0.0, step=0.01, format="%.2f")
                data_t = c6.date_input("Data da viagem", value=date.today(), key="trecho_data")

                with st.expander("Mais detalhes (opcional)"):
                    c7, c8 = st.columns(2)
                    fornecedor_t = c7.text_input("Fornecedor/companhia", key="trecho_fornecedor")
                    duracao_t = c8.number_input("Duração (horas)", min_value=0.0, step=0.5, format="%.1f")
                    c9, c10 = st.columns(2)
                    km_t = c9.number_input("Km (útil pra carro)", min_value=0.0, step=1.0)
                    data_compra_t = c10.date_input(
                        "Data da compra (se diferente de hoje)", value=None, key="trecho_data_compra",
                    )
                    obs_t = st.text_input("Observação", key="trecho_obs")

                enviar_trecho = st.form_submit_button("Adicionar trecho")

            if enviar_trecho:
                if not origem_t or not destino_t:
                    st.error("Preenche origem e destino.")
                else:
                    viagem_id = None
                    maior_ordem = 0
                    for v in (viagens_resp.data or []):
                        if v["sentido"] == sentido:
                            viagem_id = v["viagem_id"]
                            if v.get("ordem") and v["ordem"] > maior_ordem:
                                maior_ordem = v["ordem"]
                    try:
                        if viagem_id is None:
                            nova_viagem = supabase.rpc("viajai_criar_viagem", {
                                "p_folga_id": folga_id_sel, "p_sentido": sentido,
                            }).execute()
                            viagem_id = nova_viagem.data
                        supabase.rpc("viajai_adicionar_trecho", {
                            "p_viagem_id": viagem_id,
                            "p_ordem": maior_ordem + 1,
                            "p_origem": origem_t,
                            "p_destino": destino_t,
                            "p_modal": modal,
                            "p_km": km_t or None,
                            "p_data": data_t.isoformat() if data_t else None,
                            "p_preco": preco_t or None,
                            "p_fornecedor": fornecedor_t or None,
                            "p_observacao": obs_t or None,
                            "p_duracao_horas": duracao_t or None,
                            "p_data_compra": data_compra_t.isoformat() if data_compra_t else None,
                        }).execute()
                        st.success("Trecho adicionado.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao adicionar trecho: {e}")

            with st.form("form_add_gasto"):
                st.write("Adicionar gasto extra (hospedagem, alimentação, transporte local...)")
                c1, c2 = st.columns(2)
                tipo_g = c1.selectbox("Tipo", ["hospedagem", "transporte_local", "alimentacao", "outro"])
                valor_g = c2.number_input("Valor (R$)", min_value=0.0, step=0.01, format="%.2f", key="gasto_valor")
                data_g = st.date_input("Data", value=date.today(), key="gasto_data")
                obs_g = st.text_input("Observação (opcional)", key="gasto_obs")
                enviar_gasto = st.form_submit_button("Adicionar gasto")

            if enviar_gasto:
                if not valor_g:
                    st.error("Preenche o valor.")
                else:
                    try:
                        supabase.rpc("viajai_registrar_gasto", {
                            "p_folga_id": folga_id_sel,
                            "p_tipo": tipo_g,
                            "p_valor": valor_g,
                            "p_data": data_g.isoformat() if data_g else None,
                            "p_observacao": obs_g or None,
                        }).execute()
                        st.success("Gasto registrado.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao registrar gasto: {e}")

        st.divider()
        st.write("**Comparativo de custo x desvio de planejamento**")
        comp = supabase.rpc("viajai_comparativo_custo_folga", {"p_limite": 200}).execute()
        if comp.data:
            df_comp = pd.DataFrame(comp.data)
            st.dataframe(df_comp, hide_index=True, use_container_width=True)
            _botao_exportar_excel(df_comp, "viajai_comparativo_custo.xlsx")
        else:
            st.caption("Nenhum custo registrado ainda pra comparar com o desvio de planejamento.")

    with aba_rapido:
        st.caption(
            "Pra quando não dá pra apontar folga específica na hora "
            "(ex.: \"comprei 10 passagens do Ceará pra SP\") — serve pra "
            "somar gasto por rota/período e ajudar a prever."
        )
        with st.form("form_lancamento_rapido"):
            c1, c2 = st.columns(2)
            origem_lr = c1.text_input("Origem", key="lr_origem")
            destino_lr = c2.text_input("Destino", key="lr_destino")
            c3, c4, c5 = st.columns(3)
            modal_lr = c3.selectbox("Modal", ["aviao", "onibus", "carro", "taxi", "outro"], key="lr_modal")
            qtd_lr = c4.number_input("Quantidade de passagens", min_value=1, value=1, step=1, key="lr_qtd")
            valor_lr = c5.number_input("Valor total (R$)", min_value=0.0, step=0.01, format="%.2f", key="lr_valor")
            data_lr = st.date_input("Data da compra", value=date.today(), key="lr_data")
            obs_lr = st.text_input("Observação (opcional)", key="lr_obs")
            enviar_lr = st.form_submit_button("Registrar")

        if enviar_lr:
            if not origem_lr or not destino_lr or not valor_lr:
                st.error("Preenche origem, destino e valor.")
            else:
                try:
                    supabase.rpc("viajai_registrar_lancamento_rapido", {
                        "p_origem": origem_lr,
                        "p_destino": destino_lr,
                        "p_valor_total": valor_lr,
                        "p_quantidade": int(qtd_lr),
                        "p_modal": modal_lr,
                        "p_data": data_lr.isoformat() if data_lr else None,
                        "p_observacao": obs_lr or None,
                    }).execute()
                    st.success("Lançamento registrado.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Erro ao registrar lançamento: {e}")

        st.divider()
        st.write("**Lançamentos recentes**")
        lancs = supabase.rpc("viajai_listar_lancamentos_rapidos", {"p_limite": 200}).execute()
        if lancs.data:
            df_lanc = pd.DataFrame(lancs.data)
            st.dataframe(df_lanc, hide_index=True, use_container_width=True)
            _botao_exportar_excel(df_lanc, "viajai_lancamentos_rapidos.xlsx")
        else:
            st.caption("Nenhum lançamento rápido ainda.")

        st.divider()
        st.write("**Resumo de custo por rota**")
        resumo = supabase.rpc("viajai_resumo_custo_por_rota", {"p_limite": 100}).execute()
        if resumo.data:
            df_resumo = pd.DataFrame(resumo.data)
            st.dataframe(df_resumo, hide_index=True, use_container_width=True)
            _botao_exportar_excel(df_resumo, "viajai_resumo_custo_rota.xlsx")
        else:
            st.caption("Nenhum custo registrado ainda (nem lançamento rápido nem trecho com preço).")

    st.divider()
    st.write("**Gasto por período**")
    st.caption("Soma tudo que tem data (trecho + gasto extra + lançamento rápido), mês a mês.")
    meses_janela = st.slider("Últimos quantos meses?", min_value=1, max_value=24, value=12, key="periodo_meses")
    periodo_resp = supabase.rpc("viajai_gasto_por_periodo", {"p_meses": meses_janela}).execute()
    if periodo_resp.data:
        df_periodo = pd.DataFrame(periodo_resp.data)
        st.bar_chart(df_periodo.set_index("periodo")["valor_total"])
        st.dataframe(df_periodo, hide_index=True, use_container_width=True)
        _botao_exportar_excel(df_periodo, "viajai_gasto_por_periodo.xlsx")
    else:
        st.caption("Nenhum custo com data registrada ainda nesse período.")



# ---------- Assistente (chat) — mesmo padrao validado do TIA.go/app_tiago.py ----------
# (lido direto do arquivo real antes de desenhar isso, nao suposicao):
# loop de tool-use da Anthropic, 2 fases (grava pergunta + rerun, so' entao
# chama a API) pra evitar bug de ordem de mensagem, dolar escapado no
# markdown. DIFERENCA proposital do TIA.go: aqui o historico e' PERSISTIDO
# por usuario (schema_v0.11), o TIA.go so' guarda em session_state (some ao
# dar F5) - pedido explicito do Rafael (02/09): "gostaria que a memoria do
# chat fosse preservada pra ajudar nas decisoes".
#
# ESCOPO v1 (decisao 02/09, ver 00-handoff): SO CONSULTA. Nenhuma ferramenta
# de escrita ainda - registrar lancamento/trecho por chat fica pra depois,
# so' depois de validar que a parte de leitura funciona bem de verdade.

TOOLS_VIAJAI = [
    {
        "name": "consultar_previsao_folgas",
        "description": "Previsao de folga por colaborador ativo: urgencia, dias restantes, datas previstas, previsao de gasto. Use para 'quem esta de folga', 'quem precisa viajar', 'urgencia'.",
        "input_schema": {"type": "object", "properties": {}},
    },
    {
        "name": "consultar_pendencias_import",
        "description": "Linhas de import RE090 que nao bateram com nenhum colaborador do RH (pendencia de revisao manual).",
        "input_schema": {
            "type": "object",
            "properties": {
                "apenas_nao_resolvidas": {"type": "boolean", "description": "So' as ainda nao marcadas como resolvidas (padrao true)"}
            },
        },
    },
    {
        "name": "consultar_historico_folga",
        "description": "Historico de mudancas de status de folga (quem mudou, quando, de qual valor pra qual).",
        "input_schema": {
            "type": "object",
            "properties": {"limite": {"type": "integer", "description": "Quantos registros (padrao 200)"}},
        },
    },
    {
        "name": "consultar_desvio_planejamento",
        "description": "Folgas ja confirmadas/realizadas comparando data prevista x real (quantos dias atrasou/antecipou na saida e no retorno).",
        "input_schema": {
            "type": "object",
            "properties": {"limite": {"type": "integer", "description": "Quantos registros (padrao 200)"}},
        },
    },
    {
        "name": "consultar_previsao_gasto_colaborador",
        "description": "Previsao de gasto por colaborador: media do historico da propria pessoa, ou do canteiro/obra dela se nao tiver historico proprio. Devolve tambem 'base_previsao' (colaborador/canteiro/obra/sem_dado) - sempre informe essa base na resposta.",
        "input_schema": {"type": "object", "properties": {}},
    },
    {
        "name": "consultar_comparativo_custo_folga",
        "description": "Folgas com custo real ja registrado (passagem + gasto extra), cruzado com o desvio de planejamento em dias.",
        "input_schema": {
            "type": "object",
            "properties": {"limite": {"type": "integer", "description": "Quantos registros (padrao 200)"}},
        },
    },
    {
        "name": "consultar_sugestao_fornecedor_rota",
        "description": "Fornecedor/companhia mais usado historicamente numa rota especifica (origem/destino precisam bater com o texto ja registrado).",
        "input_schema": {
            "type": "object",
            "properties": {
                "origem": {"type": "string"},
                "destino": {"type": "string"},
            },
            "required": ["origem", "destino"],
        },
    },
    {
        "name": "consultar_comparar_modais_rota",
        "description": "Compara preco medio e duracao media entre modais (aviao/onibus/carro/etc) numa rota especifica, baseado no que ja foi registrado.",
        "input_schema": {
            "type": "object",
            "properties": {
                "origem": {"type": "string"},
                "destino": {"type": "string"},
            },
            "required": ["origem", "destino"],
        },
    },
    {
        "name": "consultar_resumo_custo_por_rota",
        "description": "Soma de gasto e quantidade de passagens por rota (origem/destino), juntando lancamento rapido e trecho registrado.",
        "input_schema": {
            "type": "object",
            "properties": {"limite": {"type": "integer", "description": "Quantas rotas (padrao 100)"}},
        },
    },
    {
        "name": "consultar_gasto_por_periodo",
        "description": "Soma de todo gasto registrado (trecho + gasto extra + lancamento rapido), agrupado por mes.",
        "input_schema": {
            "type": "object",
            "properties": {"meses": {"type": "integer", "description": "Quantos meses pra tras (padrao 12)"}},
        },
    },
    {
        "name": "consultar_lancamentos_rapidos",
        "description": "Lancamentos rapidos de compra registrados sem apontar folga especifica (ex.: 'comprei 10 passagens do Ceara pra SP').",
        "input_schema": {
            "type": "object",
            "properties": {"limite": {"type": "integer", "description": "Quantos registros (padrao 200)"}},
        },
    },
    {
        "name": "consultar_localizacoes_canteiro",
        "description": "Localizacao (cidade/UF/endereco/aeroporto mais proximo) dos canteiros que ja tem esse cadastro feito - nem todos tem ainda, e' alimentado aos poucos.",
        "input_schema": {"type": "object", "properties": {}},
    },
    {
        "name": "propor_lancamento_rapido",
        "description": (
            "Prepara um lancamento rapido de compra de passagem (ex.: 'comprei 10 passagens do "
            "Ceara pra SP por 3500') pra o usuario CONFIRMAR na tela antes de gravar. NUNCA grava "
            "direto no banco - so' monta a proposta, que aparece no painel lateral com um botao "
            "de confirmar. So' chame com origem, destino e valor_total certos (extraidos do que "
            "o usuario disse); se faltar algum desses 3, pergunte antes de chamar a ferramenta."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "origem": {"type": "string", "description": "Cidade/local de origem"},
                "destino": {"type": "string", "description": "Cidade/local de destino"},
                "valor_total": {"type": "number", "description": "Valor total pago (soma de todas as passagens desse lote)"},
                "quantidade": {"type": "integer", "description": "Quantas passagens (padrao 1)"},
                "modal": {"type": "string", "enum": ["aviao", "onibus", "carro", "taxi", "outro"], "description": "Padrao aviao"},
                "data": {"type": "string", "description": "Data da compra, formato YYYY-MM-DD (padrao hoje)"},
                "observacao": {"type": "string", "description": "Detalhe extra opcional (ex.: nome do fornecedor, se foi dito)"},
            },
            "required": ["origem", "destino", "valor_total"],
        },
    },
]


def _executar_ferramenta_viajai(supabase, nome, entrada):
    try:
        if nome == "consultar_previsao_folgas":
            r = supabase.rpc("viajai_previsao_folgas").execute()
        elif nome == "consultar_pendencias_import":
            r = supabase.rpc("viajai_listar_pendencias_import", {
                "p_apenas_nao_resolvidas": entrada.get("apenas_nao_resolvidas", True),
            }).execute()
        elif nome == "consultar_historico_folga":
            r = supabase.rpc("viajai_listar_historico_folga", {"p_limite": entrada.get("limite", 200)}).execute()
        elif nome == "consultar_desvio_planejamento":
            r = supabase.rpc("viajai_listar_folgas_desvio", {"p_limite": entrada.get("limite", 200)}).execute()
        elif nome == "consultar_previsao_gasto_colaborador":
            r = supabase.rpc("viajai_previsao_gasto_colaborador").execute()
        elif nome == "consultar_comparativo_custo_folga":
            r = supabase.rpc("viajai_comparativo_custo_folga", {"p_limite": entrada.get("limite", 200)}).execute()
        elif nome == "consultar_sugestao_fornecedor_rota":
            r = supabase.rpc("viajai_sugestao_fornecedor_rota", {
                "p_origem": entrada.get("origem", ""), "p_destino": entrada.get("destino", ""),
            }).execute()
        elif nome == "consultar_comparar_modais_rota":
            r = supabase.rpc("viajai_comparar_modais_rota", {
                "p_origem": entrada.get("origem", ""), "p_destino": entrada.get("destino", ""),
            }).execute()
        elif nome == "consultar_resumo_custo_por_rota":
            r = supabase.rpc("viajai_resumo_custo_por_rota", {"p_limite": entrada.get("limite", 100)}).execute()
        elif nome == "consultar_gasto_por_periodo":
            r = supabase.rpc("viajai_gasto_por_periodo", {"p_meses": entrada.get("meses", 12)}).execute()
        elif nome == "consultar_lancamentos_rapidos":
            r = supabase.rpc("viajai_listar_lancamentos_rapidos", {"p_limite": entrada.get("limite", 200)}).execute()
        elif nome == "consultar_localizacoes_canteiro":
            r = supabase.rpc("viajai_listar_localizacoes_canteiro").execute()
        elif nome == "propor_lancamento_rapido":
            # NUNCA grava aqui - so' monta a proposta pro usuario confirmar
            # na tela (col_dash em pagina_chat). O INSERT de verdade so'
            # acontece quando a pessoa clica "Confirmar e registrar", que
            # chama viajai_registrar_lancamento_rapido direto (sem passar
            # pela IA de novo) - pedido do Rafael 03/09: IA pode propor
            # acao, mas quem executa e' sempre o humano clicando.
            origem = (entrada.get("origem") or "").strip()
            destino = (entrada.get("destino") or "").strip()
            valor_total = entrada.get("valor_total")
            if not origem or not destino or not valor_total:
                return {"erro": "faltou origem, destino ou valor_total - pergunte pro usuario antes de propor de novo"}
            st.session_state.acao_pendente_viajai = {
                "tipo": "lancamento_rapido",
                "origem": origem,
                "destino": destino,
                "valor_total": float(valor_total),
                "quantidade": int(entrada.get("quantidade") or 1),
                "modal": entrada.get("modal") or "aviao",
                "data": entrada.get("data") or date.today().isoformat(),
                "observacao": entrada.get("observacao") or "",
            }
            return {
                "status": "proposta pronta, aguardando o usuario confirmar no painel lateral (nao foi gravado ainda)",
                "resumo": f"{entrada.get('quantidade') or 1}x {origem} -> {destino}, R$ {float(valor_total):.2f}",
            }
        else:
            return {"erro": "ferramenta desconhecida"}
        return r.data if r.data else []
    except Exception as e:
        return {"erro": str(e)}


def _montar_system_prompt_viajai(usuario_email):
    return (
        "Voce e o assistente do Viaj.AI, sistema de gestao de folgas/viagens/custo de "
        "colaboradores em obra da EnerMais. Ajuda a Amanda (compradora/logistica) e outras "
        f"pessoas autorizadas. A pessoa logada agora e {usuario_email}.\n\n"
        f"A data de HOJE e {date.today().isoformat()} — use esse fato pra resolver qualquer "
        "expressao de data relativa ('proximos 30 dias', 'esse mes', 'mes passado'), nunca "
        "calcule data de cabeca.\n\n"
        "Responda so com dado que veio de verdade das ferramentas — nunca invente numero, "
        "preco, rota ou nome de fornecedor. Se a ferramenta nao trouxer dado suficiente, "
        "diga isso claramente em vez de estimar ou supor.\n\n"
        "IMPORTANTE sobre preco de passagem: voce NAO tem acesso a busca de preco em tempo "
        "real (decisao de projeto, 02/09) — toda 'previsao' ou 'estimativa' de custo vem do "
        "HISTORICO ja registrado no proprio Viaj.AI, nunca do seu conhecimento geral sobre "
        "preco de passagem. Sem historico suficiente pra uma rota ou colaborador, diga isso "
        "em vez de chutar um valor plausivel.\n\n"
        "Voce PODE propor o registro de uma compra de passagem via "
        "propor_lancamento_rapido quando o usuario contar que comprou (ex.: 'comprei 10 "
        "passagens do Ceara pra SP por 3500') — mas isso NUNCA grava direto: so' monta uma "
        "proposta que aparece pro usuario confirmar na tela antes de qualquer coisa ir pro "
        "banco. So' chame essa ferramenta com origem, destino e valor_total certos; se "
        "faltar algum, pergunte antes (nunca invente valor pra completar). Fora isso, voce "
        "ainda so' CONSULTA — outras acoes (confirmar folga, registrar trecho/gasto de uma "
        "folga especifica) nao tem ferramenta ainda, oriente a usar a tela correspondente.\n\n"
        "Guia de qual ferramenta usar:\n"
        "- quem esta de folga / precisa viajar / urgencia -> consultar_previsao_folgas.\n"
        "- pendencia de import / nao bateu no RH -> consultar_pendencias_import.\n"
        "- historico de mudanca de status de folga -> consultar_historico_folga.\n"
        "- quem atrasou / desvio do planejado -> consultar_desvio_planejamento.\n"
        "- previsao de gasto por colaborador -> consultar_previsao_gasto_colaborador "
        "(sempre informe o 'base_previsao' que vier na resposta).\n"
        "- custo real x desvio de uma folga -> consultar_comparativo_custo_folga.\n"
        "- qual companhia mais usamos numa rota -> consultar_sugestao_fornecedor_rota "
        "(precisa origem e destino).\n"
        "- aviao x onibus x carro numa rota -> consultar_comparar_modais_rota (precisa "
        "origem e destino).\n"
        "- gasto total por rota -> consultar_resumo_custo_por_rota.\n"
        "- gasto por mes/periodo -> consultar_gasto_por_periodo.\n"
        "- lancamentos rapidos recentes -> consultar_lancamentos_rapidos.\n"
        "- onde fica um canteiro / aeroporto mais proximo -> consultar_localizacoes_canteiro "
        "(nem todo canteiro tem cadastro ainda, avise se nao achar).\n"
        "- usuario disse que COMPROU passagem(ns) -> propor_lancamento_rapido (so' com "
        "origem/destino/valor certos; nunca grava sozinho, so' propoe pro usuario confirmar)."
    )


FERRAMENTAS_VISUAIS_VIAJAI = {tool["name"] for tool in TOOLS_VIAJAI if tool["name"] != "propor_lancamento_rapido"}
# Toda ferramenta de consulta pode virar tabela no painel lateral do chat;
# a de localizacao de canteiro tambem tenta virar mapa (se tiver lat/lon
# cadastrada) — pedido do Rafael 02/09: "gerar um dash interativo ao lado,
# mostrando tabela, mapa". Mesmo padrao do TIA.go (col_dash / dash_extra),
# adaptado pras ferramentas do Viaj.AI.


def pagina_chat(supabase):
    st.subheader("Assistente Viaj.AI")
    st.caption(
        "Converse em português sobre folgas, urgência, custo e histórico — só responde com "
        "dado real do Viaj.AI (nunca busca preço na internet nem inventa número). Já entende "
        "'comprei N passagens de X pra Y' e prepara o lançamento, mas só grava depois que "
        "você confirmar no painel ao lado — nunca escreve sozinho."
    )

    if not ANTHROPIC_API_KEY:
        st.warning(
            "Chave da Anthropic ainda não configurada (Secrets do Streamlit Cloud: "
            "ANTHROPIC_API_KEY) — o assistente não funciona sem ela."
        )
        return

    if "mensagens_chat" not in st.session_state:
        hist = supabase.rpc("viajai_listar_historico_chat", {"p_limite": 50}).execute()
        st.session_state.mensagens_chat = [
            {"role": m["papel"], "content": m["conteudo"]} for m in (hist.data or [])
        ]

    def _escapar_dolar(texto):
        return texto.replace("$", "\\$")

    col_chat, col_dash = st.columns([3, 2])

    with col_chat:
        for m in st.session_state.mensagens_chat:
            with st.chat_message(m["role"]):
                conteudo = m["content"] if isinstance(m["content"], str) else "(ferramenta)"
                st.markdown(_escapar_dolar(conteudo))

        pergunta = st.chat_input("Pergunte sobre folgas, urgência, custo...")

        # Mesmo padrao 2 fases do TIA.go (evita bug de ordem visto la ja em
        # producao): grava a pergunta e recarrega antes de chamar a API.
        if pergunta:
            st.session_state.mensagens_chat.append({"role": "user", "content": pergunta})
            supabase.rpc("viajai_salvar_mensagem_chat", {"p_papel": "user", "p_conteudo": pergunta}).execute()
            st.rerun()

        if st.session_state.mensagens_chat and st.session_state.mensagens_chat[-1]["role"] == "user":
            with st.spinner("Consultando..."):
                texto_final = None
                try:
                    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
                    mensagens_api = [
                        {"role": m["role"], "content": m["content"]}
                        for m in st.session_state.mensagens_chat
                        if isinstance(m["content"], str)
                    ]
                    system_prompt = _montar_system_prompt_viajai(st.session_state.usuario)

                    resposta = client.messages.create(
                        model=MODEL_ID, max_tokens=4096, system=system_prompt,
                        tools=TOOLS_VIAJAI, messages=mensagens_api,
                    )
                    # Limite de voltas de ferramenta (protecao contra loop
                    # longo demais em pergunta com muito calculo/data
                    # encadeado - achado 02/09: pergunta complexa de
                    # previsao ficou sem resposta nenhuma, ver nota abaixo).
                    _voltas_ferramenta = 0
                    while resposta.stop_reason == "tool_use" and _voltas_ferramenta < 8:
                        _voltas_ferramenta += 1
                        tool_uses = [b for b in resposta.content if b.type == "tool_use"]
                        resultados = []
                        for tu in tool_uses:
                            resultado = _executar_ferramenta_viajai(supabase, tu.name, tu.input)
                            resultados.append(
                                {"type": "tool_result", "tool_use_id": tu.id, "content": str(resultado)}
                            )
                            if tu.name in FERRAMENTAS_VISUAIS_VIAJAI:
                                st.session_state.dash_extra_viajai = {
                                    "tool": tu.name,
                                    "input": tu.input,
                                    "resultado": resultado,
                                }
                        mensagens_api.append({"role": "assistant", "content": resposta.content})
                        mensagens_api.append({"role": "user", "content": resultados})
                        resposta = client.messages.create(
                            model=MODEL_ID, max_tokens=4096, system=system_prompt,
                            tools=TOOLS_VIAJAI, messages=mensagens_api,
                        )
                    texto_final = "".join(b.text for b in resposta.content if b.type == "text")
                    # Achado 02/09: pergunta com bastante calculo (varias
                    # ferramentas + conta de data encadeada) as vezes cortava
                    # a resposta sem nenhum texto (stop_reason vinha
                    # "max_tokens" bem no meio de uma chamada de ferramenta,
                    # o while acima nao pega isso porque so continua se
                    # stop_reason == "tool_use") - bolha vazia no chat, sem
                    # erro nenhum pra avisar o que houve. max_tokens subiu
                    # de 1024 pra 4096 (reduz bastante a chance) e, se ainda
                    # assim vier vazio, avisa em vez de ficar mudo.
                    if not texto_final.strip():
                        if resposta.stop_reason == "max_tokens":
                            texto_final = (
                                "A resposta ficou grande demais e foi cortada antes de terminar "
                                "(pergunta com bastante calculo/etapa junto). Tenta perguntar em "
                                "partes menores ou de um jeito mais direto."
                            )
                        elif _voltas_ferramenta >= 8:
                            texto_final = (
                                "Essa pergunta precisou de muitas consultas em sequencia e eu parei "
                                "antes de concluir, pra não travar. Tenta quebrar em perguntas menores."
                            )
                        else:
                            texto_final = "Não consegui gerar uma resposta pra essa pergunta — tenta reformular."
                except Exception as e:
                    texto_final = f"Erro ao consultar o assistente: {e}"

            st.session_state.mensagens_chat.append({"role": "assistant", "content": texto_final})
            supabase.rpc("viajai_salvar_mensagem_chat", {"p_papel": "assistant", "p_conteudo": texto_final}).execute()
            st.rerun()

    with col_dash:
        acao = st.session_state.get("acao_pendente_viajai")
        if acao and acao.get("tipo") == "lancamento_rapido":
            # Confirmacao humana antes de gravar - pedido do Rafael 03/09:
            # a IA so' PROPOE (ferramenta propor_lancamento_rapido nunca
            # grava), quem executa de verdade e' sempre a pessoa clicando
            # aqui. Campos vem preenchidos com o que a IA extraiu, mas dao
            # pra corrigir antes de confirmar - camada extra de seguranca.
            st.warning("Confirme antes de gravar — extraído da conversa, revise se estiver errado")
            with st.form("form_confirmar_lancamento_ia"):
                c_origem = st.text_input("Origem", value=acao["origem"])
                c_destino = st.text_input("Destino", value=acao["destino"])
                c_valor = st.number_input("Valor total (R$)", value=float(acao["valor_total"]), min_value=0.0, step=0.01)
                c_qtd = st.number_input("Quantidade", value=int(acao["quantidade"]), min_value=1, step=1)
                c_modal = st.selectbox(
                    "Modal", ["aviao", "onibus", "carro", "taxi", "outro"],
                    index=["aviao", "onibus", "carro", "taxi", "outro"].index(acao["modal"]) if acao["modal"] in ["aviao", "onibus", "carro", "taxi", "outro"] else 0,
                )
                try:
                    _data_default = date.fromisoformat(acao["data"])
                except (ValueError, TypeError):
                    _data_default = date.today()
                c_data = st.date_input("Data", value=_data_default)
                c_obs = st.text_input("Observação", value=acao.get("observacao", ""))
                col_ok, col_no = st.columns(2)
                confirmar = col_ok.form_submit_button("✅ Confirmar e registrar", use_container_width=True)
                cancelar = col_no.form_submit_button("Cancelar", use_container_width=True)

            if confirmar:
                supabase.rpc("viajai_registrar_lancamento_rapido", {
                    "p_origem": c_origem, "p_destino": c_destino, "p_valor_total": c_valor,
                    "p_quantidade": int(c_qtd), "p_modal": c_modal,
                    "p_data": c_data.isoformat(), "p_observacao": c_obs or None,
                }).execute()
                del st.session_state.acao_pendente_viajai
                st.session_state.mensagens_chat.append({
                    "role": "assistant",
                    "content": f"Registrado: {int(c_qtd)}x {c_origem} -> {c_destino}, R$ {c_valor:.2f}.",
                })
                supabase.rpc("viajai_salvar_mensagem_chat", {
                    "p_papel": "assistant",
                    "p_conteudo": f"Registrado: {int(c_qtd)}x {c_origem} -> {c_destino}, R$ {c_valor:.2f}.",
                }).execute()
                st.success("Registrado.")
                st.rerun()
            elif cancelar:
                del st.session_state.acao_pendente_viajai
                st.rerun()
            st.divider()

        st.caption("Painel — última consulta com dado")
        extra = st.session_state.get("dash_extra_viajai")
        if not extra:
            st.info(
                "Faça uma pergunta que puxe dado (folga, custo, canteiro...) pra ver a "
                "tabela (e mapa, se for canteiro com localização) aqui."
            )
        else:
            resultado = extra.get("resultado")
            if not isinstance(resultado, list) or not resultado:
                st.write(resultado if resultado else "(sem dado pra essa consulta)")
            else:
                df_dash = pd.DataFrame(resultado)
                st.dataframe(df_dash, use_container_width=True, hide_index=True)
                if extra.get("tool") == "consultar_localizacoes_canteiro" and {
                    "latitude", "longitude"
                }.issubset(df_dash.columns):
                    df_mapa = df_dash.dropna(subset=["latitude", "longitude"])
                    if not df_mapa.empty:
                        st.map(df_mapa, latitude="latitude", longitude="longitude", size=15)
                    else:
                        st.caption("Nenhum canteiro com latitude/longitude cadastrada ainda.")

def main():
    if "sessao" not in st.session_state:
        tela_login()
        return

    # Reanexa o token a cada rerun (padrão TIA.go — Streamlit recria o
    # cliente do zero a cada interação, sem isso a consulta vai como anônimo
    # e a RLS devolve vazio sem erro nenhum).
    supabase = get_client()
    supabase.postgrest.auth(st.session_state.sessao.access_token)

    with st.sidebar:
        render_logo(height=56)
        st.write(f"Logado como: {st.session_state.usuario}")
        pagina = st.radio(
            "Navegação",
            ["Importar RE090", "Confirmar folgas", "Previsão de folgas", "Custo & Passagens", "Assistente"],
        )
        if st.button("Sair"):
            supabase.auth.sign_out()
            del st.session_state.sessao
            st.rerun()

    if pagina == "Importar RE090":
        pagina_importar_re090(supabase)
    elif pagina == "Confirmar folgas":
        pagina_confirmar_folgas(supabase)
    elif pagina == "Previsão de folgas":
        pagina_previsao(supabase)
    elif pagina == "Custo & Passagens":
        pagina_custo_passagens(supabase)
    elif pagina == "Assistente":
        pagina_chat(supabase)


if __name__ == "__main__":
    main()
